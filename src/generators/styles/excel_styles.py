"""Estilos openpyxl reutilizables para el Excel del mesociclo.

Funciones puras que devuelven objetos de estilo nuevos en cada llamada
(openpyxl no permite compartir el mismo objeto entre celdas).
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.generators.styles.colors import (
    BG_HEADER,
    COLOR_PRIMARY,
    COLOR_SECONDARY,
    COLOR_WHITE,
    argb,
)

_THIN: Side = Side(style="thin", color="FFBDBDBD")


def thin_border() -> Border:
    """Borde fino gris en los cuatro lados."""
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def fill(hex_color: str) -> PatternFill:
    """Relleno sólido a partir de un color `#RRGGBB`."""
    code: str = argb(hex_color)
    return PatternFill(start_color=code, end_color=code, fill_type="solid")


def title_font() -> Font:
    """Fuente del título principal (16pt bold)."""
    return Font(name="Calibri", size=16, bold=True, color=argb(COLOR_PRIMARY))


def subtitle_font() -> Font:
    """Fuente del subtítulo (11pt, gris)."""
    return Font(name="Calibri", size=11, color=argb(COLOR_SECONDARY))


def header_font() -> Font:
    """Fuente de cabecera de tabla (blanca, bold)."""
    return Font(name="Calibri", size=11, bold=True, color=argb(COLOR_WHITE))


def micro_header_font() -> Font:
    """Fuente de cabecera de columna de microciclo (texto oscuro, bold)."""
    return Font(name="Calibri", size=11, bold=True, color=argb(COLOR_PRIMARY))


def day_font() -> Font:
    """Fuente de la fila separadora de día (blanca, bold)."""
    return Font(name="Calibri", size=12, bold=True, color=argb(COLOR_WHITE))


def body_font() -> Font:
    """Fuente del cuerpo de la tabla."""
    return Font(name="Calibri", size=11, color=argb(COLOR_PRIMARY))


def section_font() -> Font:
    """Fuente de título de sección en la hoja de notas."""
    return Font(name="Calibri", size=13, bold=True, color=argb(COLOR_PRIMARY))


def header_fill() -> PatternFill:
    """Relleno de la cabecera de tabla."""
    return fill(BG_HEADER)


def centered(*, wrap: bool = False) -> Alignment:
    """Alineación centrada (opcionalmente con ajuste de texto)."""
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def left(*, wrap: bool = False) -> Alignment:
    """Alineación a la izquierda (opcionalmente con ajuste de texto)."""
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)
