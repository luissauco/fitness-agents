"""Generador del Excel de mesociclo.

Replica el formato de referencia: una hoja con el programa de
entrenamiento (un día por bloque, una columna por microciclo para que el
usuario rellene KGs×Reps en el gym), una hoja de esquema semanal y una
hoja de notas/técnicas.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.generators.base import FileGenerator
from src.generators.styles import excel_styles as st
from src.generators.styles.colors import (
    BG_REST,
    COLOR_ACCENT,
    microcycle_color,
)
from src.models.mesocycle import Mesocycle, Microcycle, SetScheme, TrainingDay

_logger: Final[logging.Logger] = logging.getLogger(__name__)

# Explicaciones de cada técnica avanzada para la hoja de notas.
_TECHNIQUE_EXPLANATIONS: Final[dict[str, str]] = {
    "top_back_off": (
        "Top set / back-off: una serie pesada de pocas reps (top) seguida de "
        "series más ligeras y con más reps (back-off) ajustando el peso."
    ),
    "rest_pause": (
        "Rest-pause: tras llegar al fallo, descansa 15-20s y haz tantas reps "
        "como puedas. Repite 1-2 veces."
    ),
    "drop_set": (
        "Drop set: tras la última serie al fallo, baja 20-30% del peso y haz más reps al fallo."
    ),
    "superset": (
        "Superserie: dos ejercicios encadenados sin descanso entre ellos "
        "(habitualmente de músculos antagonistas)."
    ),
    "myo_reps": (
        "Myo-reps: serie de activación al fallo y mini-series posteriores con "
        "descansos muy cortos para acumular reps efectivas."
    ),
}


class MesocycleExcelGenerator(FileGenerator):
    """Genera el archivo Excel a partir de un `Mesocycle`."""

    def generate(self, mesocycle: Mesocycle, user_name: str) -> Path:  # type: ignore[override]
        """Genera el Excel de 3 hojas y devuelve su `Path`."""
        if not mesocycle.microcycles:
            raise ValueError("El mesociclo no tiene microciclos; no se puede generar el Excel.")

        wb: Workbook = Workbook()

        ws_main: Worksheet = wb.active
        ws_main.title = "Mesociclo"
        self._build_main_sheet(ws_main, mesocycle)

        self._build_schedule_sheet(wb.create_sheet("Esquema semanal"), mesocycle)
        self._build_notes_sheet(wb.create_sheet("Notas y técnicas"), mesocycle)

        filename: Path = self._build_filename(
            prefix="Mesociclo", identifier=user_name, extension="xlsx"
        )
        wb.save(filename)
        _logger.info("Excel de mesociclo generado: %s", filename)
        return filename

    # ----------------------------- Hoja principal -----------------------------

    def _build_main_sheet(self, ws: Worksheet, mesocycle: Mesocycle) -> None:
        """Construye la hoja con el programa y una columna por microciclo."""
        micros: list[Microcycle] = mesocycle.microcycles
        n_micros: int = len(micros)
        total_cols: int = 3 + n_micros  # A, B, micros..., Descanso
        last_col_letter: str = get_column_letter(total_cols)

        # Filas 1-2: cabecera grande.
        ws.merge_cells(f"A1:{last_col_letter}1")
        c = ws["A1"]
        c.value = f"MESOCICLO: {mesocycle.name}"
        c.font = st.title_font()
        c.alignment = st.left()

        ws.merge_cells(f"A2:{last_col_letter}2")
        c = ws["A2"]
        c.value = (
            f"{mesocycle.phase} | Inicio: {mesocycle.start_date.isoformat()} "
            f"| Split: {mesocycle.split_type} | {mesocycle.training_days_per_week} días/sem"
        )
        c.font = st.subtitle_font()
        c.alignment = st.left()

        # Fila 4: cabecera de columnas.
        header_row: int = 4
        self._write_header(ws, header_row, micros, total_cols)

        # Filas de datos: días y ejercicios del primer microciclo (estructura canónica).
        row: int = header_row + 1
        for day in micros[0].training_days:
            row = self._write_day_block(ws, row, day, total_cols)

        # Anchos de columna.
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        for i in range(n_micros):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
        ws.column_dimensions[last_col_letter].width = 10

        # Congelar cabecera (4 filas) y columnas A-B.
        ws.freeze_panes = "C5"

    def _write_header(
        self, ws: Worksheet, row: int, micros: list[Microcycle], total_cols: int
    ) -> None:
        """Escribe la fila de cabecera con una columna por microciclo."""
        border = st.thin_border()

        a = ws.cell(row=row, column=1, value="Ejercicio")
        b = ws.cell(row=row, column=2, value="Sets×Reps(RIR)")
        for cell in (a, b):
            cell.font = st.header_font()
            cell.fill = st.header_fill()
            cell.alignment = st.centered()
            cell.border = border

        n_micros: int = len(micros)
        for i, micro in enumerate(micros):
            label: str = "DESCARGA" if micro.is_deload else f"MICRO {micro.number}"
            cell = ws.cell(row=row, column=3 + i, value=label)
            cell.font = st.micro_header_font()
            cell.fill = st.fill(microcycle_color(i, n_micros, is_deload=micro.is_deload))
            cell.alignment = st.centered()
            cell.border = border

        rest = ws.cell(row=row, column=total_cols, value="Descanso")
        rest.font = st.header_font()
        rest.fill = st.header_fill()
        rest.alignment = st.centered()
        rest.border = border

    def _write_day_block(self, ws: Worksheet, row: int, day: TrainingDay, total_cols: int) -> int:
        """Escribe la fila separadora del día y sus ejercicios. Devuelve la fila siguiente."""
        last_col_letter: str = get_column_letter(total_cols)
        border = st.thin_border()

        # Fila separadora del día.
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        sep = ws.cell(row=row, column=1, value=day.day_label)
        if day.is_rest_day:
            sep.fill = st.fill(BG_REST)
            sep.font = st.body_font()
        else:
            sep.fill = st.header_fill()
            sep.font = st.day_font()
        sep.alignment = st.left()
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).border = border
        row += 1

        if day.is_rest_day:
            return row

        for exercise in sorted(day.exercises, key=lambda e: e.order):
            name: str = exercise.exercise_name
            if exercise.progression_notes:
                name = f"{name}\n({exercise.progression_notes})"

            cell = ws.cell(row=row, column=1, value=name)
            cell.font = st.body_font()
            cell.alignment = st.left(wrap=True)
            cell.border = border

            scheme: SetScheme = exercise.set_scheme
            sc = ws.cell(row=row, column=2, value=scheme.description)
            sc.font = st.body_font()
            sc.alignment = st.left(wrap=True)
            sc.border = border

            # Columnas de microciclos: vacías, color por microciclo.
            n_micros: int = total_cols - 3
            for i in range(n_micros):
                mc = ws.cell(row=row, column=3 + i)
                mc.fill = st.fill(microcycle_color(i, n_micros, is_deload=False))
                mc.border = border
                mc.alignment = st.centered()

            rest_cell = ws.cell(
                row=row, column=total_cols, value=self._format_rest(scheme.rest_seconds)
            )
            rest_cell.font = st.body_font()
            rest_cell.alignment = st.centered()
            rest_cell.border = border

            # Altura adaptativa si el nombre lleva nota en segunda línea.
            ws.row_dimensions[row].height = 32 if "\n" in name else 20
            row += 1

        return row

    @staticmethod
    def _format_rest(seconds: int) -> str:
        """Formatea el descanso como `m'ss''` (ej: 150 → 2'30'')."""
        return f"{seconds // 60}'{seconds % 60:02d}''"

    # --------------------------- Hoja esquema semanal --------------------------

    def _build_schedule_sheet(self, ws: Worksheet, mesocycle: Mesocycle) -> None:
        """Construye la hoja de esquema semanal a partir de `weekly_schedule.days`."""
        border = st.thin_border()

        ws.merge_cells("A1:C1")
        title = ws["A1"]
        title.value = "ESQUEMA SEMANAL"
        title.font = st.section_font()
        title.alignment = st.left()

        headers: list[str] = ["Día", "Tipo", "Pasos mínimos"]
        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=text)
            cell.font = st.header_font()
            cell.fill = st.header_fill()
            cell.alignment = st.centered()
            cell.border = border

        row: int = 4
        for entry in mesocycle.weekly_schedule.days:
            day_val = entry.get("day", "")
            type_val = str(entry.get("type", "")).strip()
            steps_val = entry.get("steps", "")

            is_rest: bool = type_val.lower() in {"descanso", "rest"}
            bg: str = BG_REST if is_rest else "#E7F4EE"

            for col, value in enumerate((day_val, type_val.capitalize(), steps_val), start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = st.body_font()
                cell.alignment = st.centered()
                cell.fill = st.fill(bg)
                cell.border = border
            row += 1

        note: str = (
            mesocycle.weekly_schedule.notes
            or "Este esquema es referencial. Los pasos indicados son un mínimo."
        )
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=3)
        n = ws.cell(row=row + 1, column=1, value=note)
        n.font = st.subtitle_font()
        n.alignment = st.left(wrap=True)

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16

    # ---------------------------- Hoja notas/técnicas --------------------------

    def _build_notes_sheet(self, ws: Worksheet, mesocycle: Mesocycle) -> None:
        """Construye la hoja de notación, técnicas y progresión."""
        ws.column_dimensions["A"].width = 110
        row: int = 1

        def section(titulo: str) -> None:
            nonlocal row
            cell = ws.cell(row=row, column=1, value=titulo)
            cell.fill = st.fill(COLOR_ACCENT)
            cell.font = st.day_font()
            row += 1

        def line(texto: str) -> None:
            nonlocal row
            cell = ws.cell(row=row, column=1, value=texto)
            cell.font = st.body_font()
            cell.alignment = st.left(wrap=True)
            ws.row_dimensions[row].height = 30
            row += 1

        section("NOTACIÓN DE SERIES")
        line('"RIR" = Reps In Reserve. RIR 1 significa parar a una repetición del fallo.')
        line(
            'Ej: "top set: 1x6(1) / back-off: 2x10(1)" = 1 serie top de 6 reps con RIR 1, '
            "seguida de 2 series back-off de 10 reps con RIR 1, ajustando el peso."
        )
        row += 1

        # Técnicas realmente usadas en este mesociclo.
        used: list[str] = self._collect_techniques(mesocycle)
        if used:
            section("TÉCNICAS DE INTENSIFICACIÓN")
            for tech in used:
                explanation: str | None = _TECHNIQUE_EXPLANATIONS.get(tech)
                if explanation:
                    line(f"• {explanation}")
            row += 1

        section("PROGRESIÓN")
        line(mesocycle.progression_strategy)
        if mesocycle.notes:
            row += 1
            section("INDICACIONES GENERALES")
            line(mesocycle.notes)

    @staticmethod
    def _collect_techniques(mesocycle: Mesocycle) -> list[str]:
        """Devuelve las técnicas avanzadas presentes en el mesociclo (orden estable)."""
        seen: dict[str, None] = {}
        for micro in mesocycle.microcycles:
            for day in micro.training_days:
                for ex in day.exercises:
                    tech = ex.set_scheme.technique
                    if tech and tech != "straight":
                        seen.setdefault(tech, None)
        return list(seen)
