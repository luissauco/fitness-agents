"""Tests del generador Excel de mesociclo."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.generators.xlsx_mesocycle import MesocycleExcelGenerator
from src.models.mesocycle import Mesocycle


@pytest.fixture
def generated(tmp_path: Path, mesocycle: Mesocycle) -> Path:
    gen = MesocycleExcelGenerator(output_dir=tmp_path)
    return gen.generate(mesocycle, user_name="Luis Sauco")


def test_file_created(generated: Path) -> None:
    assert generated.exists()
    assert generated.suffix == ".xlsx"
    assert "Luis_Sauco" in generated.name


def test_sheets_present(generated: Path) -> None:
    wb = load_workbook(generated)
    assert wb.sheetnames == ["Mesociclo", "Esquema semanal", "Notas y técnicas"]


def test_main_header_and_micro_columns(generated: Path, mesocycle: Mesocycle) -> None:
    wb = load_workbook(generated)
    ws = wb["Mesociclo"]

    assert ws["A4"].value == "Ejercicio"
    assert ws["B4"].value == "Sets×Reps(RIR)"

    n = len(mesocycle.microcycles)
    # Columnas C..(C+n-1) son los microciclos; la última columna es "Descanso".
    micro_labels = [ws.cell(row=4, column=3 + i).value for i in range(n)]
    assert micro_labels == ["MICRO 1", "MICRO 2", "MICRO 3", "DESCARGA"]
    assert ws.cell(row=4, column=3 + n).value == "Descanso"


def test_micro_cells_are_empty(generated: Path, mesocycle: Mesocycle) -> None:
    wb = load_workbook(generated)
    ws = wb["Mesociclo"]
    n = len(mesocycle.microcycles)

    # Localiza la primera fila de ejercicio (col A no vacía, col B con esquema).
    exercise_rows = [
        r
        for r in range(5, ws.max_row + 1)
        if ws.cell(row=r, column=1).value and ws.cell(row=r, column=2).value
    ]
    assert exercise_rows, "No se encontraron filas de ejercicio."
    for r in exercise_rows:
        for i in range(n):
            assert ws.cell(row=r, column=3 + i).value is None


def test_rest_day_row_present(generated: Path) -> None:
    wb = load_workbook(generated)
    ws = wb["Mesociclo"]
    texts = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
    assert any(t and "Descanso" in str(t) for t in texts)


def test_freeze_panes(generated: Path) -> None:
    wb = load_workbook(generated)
    assert wb["Mesociclo"].freeze_panes == "C5"


def test_schedule_sheet_has_rows(generated: Path, mesocycle: Mesocycle) -> None:
    wb = load_workbook(generated)
    ws = wb["Esquema semanal"]
    assert ws["A3"].value == "Día"
    data_rows = [
        r for r in range(4, ws.max_row + 1) if ws.cell(row=r, column=1).value not in (None, "")
    ]
    assert len(data_rows) >= len(mesocycle.weekly_schedule.days)


def test_notes_sheet_contains_sections(generated: Path) -> None:
    wb = load_workbook(generated)
    ws = wb["Notas y técnicas"]
    all_text = " ".join(str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1))
    assert "NOTACIÓN DE SERIES" in all_text
    assert "PROGRESIÓN" in all_text
    # El mesociclo de prueba usa top_back_off → debe explicarse.
    assert "Top set" in all_text


def test_empty_microcycles_raises(tmp_path: Path, mesocycle: Mesocycle) -> None:
    mesocycle.microcycles.clear()
    gen = MesocycleExcelGenerator(output_dir=tmp_path)
    with pytest.raises(ValueError, match="microciclos"):
        gen.generate(mesocycle, user_name="X")
