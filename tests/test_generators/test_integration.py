"""Test de integración end-to-end de los generadores (sin LLM).

Usa fixtures estáticos (mesocycle, nutrition_plan, progress_log) para
simular el flujo completo: tras correr los agentes, el sistema produce el
Excel del mesociclo, el PDF nutricional y el PDF de progreso para el
mismo usuario, con naming consistente y archivos válidos.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from src.generators.pdf_nutrition import NutritionPDFGenerator
from src.generators.pdf_progress import ProgressPDFGenerator
from src.generators.xlsx_mesocycle import MesocycleExcelGenerator
from src.models.mesocycle import Mesocycle
from src.models.nutrition_plan import NutritionPlan
from src.models.progress_log import ProgressLog

_USER = "Luis Sauco"
_TODAY = date.today().strftime("%Y-%m-%d")


def test_full_pipeline_generates_all_artifacts(
    tmp_path: Path,
    mesocycle: Mesocycle,
    nutrition_plan: NutritionPlan,
    progress_log: ProgressLog,
) -> None:
    """Los tres generadores comparten output_dir y producen archivos válidos."""
    out = tmp_path / "output"

    xlsx = MesocycleExcelGenerator(output_dir=out).generate(mesocycle, _USER)
    pdf_nut = NutritionPDFGenerator(output_dir=out).generate(nutrition_plan, _USER)
    pdf_prog = ProgressPDFGenerator(output_dir=out).generate(progress_log, _USER)

    # Todos existen en el mismo directorio.
    for path in (xlsx, pdf_nut, pdf_prog):
        assert path.exists()
        assert path.parent == out

    # Naming consistente: {Tipo}_{Nombre}_{fecha}.{ext}
    assert xlsx.name == f"Mesociclo_Luis_Sauco_{_TODAY}.xlsx"
    assert pdf_nut.name == f"Plan_Nutricional_Luis_Sauco_{_TODAY}.pdf"
    assert pdf_prog.name == f"Informe_Progreso_Luis_Sauco_{_TODAY}.pdf"

    # Los archivos son abribles por sus lectores.
    wb = load_workbook(xlsx)
    assert wb.sheetnames == ["Mesociclo", "Esquema semanal", "Notas y técnicas"]
    assert len(PdfReader(str(pdf_nut)).pages) >= 6
    assert len(PdfReader(str(pdf_prog)).pages) == 6

    # output/ contiene exactamente los 3 artefactos.
    assert sorted(p.name for p in out.iterdir()) == sorted([xlsx.name, pdf_nut.name, pdf_prog.name])


def test_regeneration_is_idempotent_same_day(tmp_path: Path, mesocycle: Mesocycle) -> None:
    """Regenerar el mismo día sobreescribe el archivo (mismo nombre)."""
    gen = MesocycleExcelGenerator(output_dir=tmp_path)
    first = gen.generate(mesocycle, _USER)
    second = gen.generate(mesocycle, _USER)
    assert first == second
    assert list(tmp_path.glob("*.xlsx")) == [first]


def test_progress_pdf_with_history_embeds_chart(tmp_path: Path, progress_log: ProgressLog) -> None:
    """Con histórico previo, el informe de progreso embebe la gráfica de peso."""
    prev = progress_log.model_copy(deep=True)
    prev.period_end = date(2026, 5, 5)

    path = ProgressPDFGenerator(output_dir=tmp_path).generate(
        progress_log, _USER, previous_logs=[prev]
    )
    page2 = PdfReader(str(path)).pages[1]
    assert "/XObject" in page2["/Resources"]
